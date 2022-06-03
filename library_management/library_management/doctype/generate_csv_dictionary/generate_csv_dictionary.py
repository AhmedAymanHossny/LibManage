# Copyright (c) 2022, Faris Ansari and contributors
# For license information, please see license.txt


import frappe
from frappe.model.document import Document
import openpyxl as xl


class GenerateCSVDictionary(Document):
    @frappe.whitelist()
    def get_dict(self):
        dic = {}
        SOURCE = './library.test' + \
            frappe.db.get_value('Generate CSV Dictionary',
                                self.name, 'xl_file')
        wb = loadItems(SOURCE)
        dic = Fill_Dict(self, wb)
        keys = dic.keys()
        self.dictionary = ""
        for k in keys:
            self.dictionary += (str(k) + ' ' + dic[k] + '\n')

    @frappe.whitelist()
    def validate_file(self):
        s = frappe.db.get_value(
            'Generate CSV Dictionary', self.name, 'xl_file')
        try:
            if (s.split(".")[1] == 'xlsx'):
                return True
            else:
                frappe.throw('Invalid File Type!')
        except:
            return False

    @frappe.whitelist()
    def get_sheet_name(self):
        SOURCE = './library.test' + \
            frappe.db.get_value('Generate CSV Dictionary',
                                self.name, 'xl_file')
        wb = loadItems(SOURCE)
        sheets = wb.sheetnames
        return sheets


def loadItems(SOURCE):
    wb = xl.load_workbook(SOURCE)
    return wb


def Fill_Dict(self, wb):
    dict = {}
    lst = []
    itr = 1
    try:
        sheet = wb[self.sheet_title]
        for row in range(2, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, col)
                lst.append(cell.value)
            dict[itr] = str(lst)
            lst.clear()
            itr += 1
    except:
        frappe.throw("Select Sheet First!")
    return dict


"""
def Fill_Dict(wb):
    dict = {}
    lst = []
    itr = 1
    for sheet in wb.worksheets:
        for row in range(2, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, col)
                lst.append(cell.value)
            dict[itr] = str(lst)
            lst.clear()
            itr += 1
    return dict
"""
